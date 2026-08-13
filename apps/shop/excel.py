# apps/shop/excel.py
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.shop.models import Order


def export_orders_excel(queryset):
    """
    Export selected orders to Excel.
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "سفارشات"

    # ===== هدرها =====
    headers = [
        'ردیف',
        'شماره سفارش',
        'مشتری',
        'شماره موبایل',
        'آدرس',
        'تعداد محصولات',
        'مبلغ کل',
        'وضعیت سفارش',
        'وضعیت پرداخت',
        'تاریخ ثبت',
    ]

    # استایل هدر
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F6FED', end_color='2F6FED', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ===== داده‌ها =====
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    for row, order in enumerate(queryset, start=2):
        # شماره ردیف
        worksheet.cell(row=row, column=1).value = row - 1
        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        # شماره سفارش
        worksheet.cell(row=row, column=2).value = order.order_number
        worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        # مشتری
        customer_name = order.user.fullname or order.user.username
        worksheet.cell(row=row, column=3).value = customer_name

        # شماره موبایل
        worksheet.cell(row=row, column=4).value = order.user.mobile or '-'
        worksheet.cell(row=row, column=4).alignment = Alignment(horizontal='center')

        # آدرس
        address = order.address.address if order.address else '-'
        worksheet.cell(row=row, column=5).value = address
        worksheet.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

        # تعداد محصولات
        total_items = order.items.aggregate(total=Sum('quantity'))['total'] or 0
        worksheet.cell(row=row, column=6).value = total_items
        worksheet.cell(row=row, column=6).alignment = Alignment(horizontal='center')

        # مبلغ کل
        worksheet.cell(row=row, column=7).value = float(order.total)
        worksheet.cell(row=row, column=7).number_format = '#,##0'
        worksheet.cell(row=row, column=7).alignment = Alignment(horizontal='right')

        # وضعیت سفارش
        status_label = dict(Order.STATUS_CHOICES).get(order.status, order.status)
        worksheet.cell(row=row, column=8).value = status_label
        worksheet.cell(row=row, column=8).alignment = Alignment(horizontal='center')

        # وضعیت پرداخت
        payment_label = dict(Order.PAYMENT_STATUS_CHOICES).get(order.payment_status, order.payment_status)
        worksheet.cell(row=row, column=9).value = payment_label
        worksheet.cell(row=row, column=9).alignment = Alignment(horizontal='center')

        # تاریخ ثبت
        worksheet.cell(row=row, column=10).value = order.created_at.strftime('%Y/%m/%d %H:%M')
        worksheet.cell(row=row, column=10).alignment = Alignment(horizontal='center')

        # اعمال border
        for col in range(1, len(headers) + 1):
            worksheet.cell(row=row, column=col).border = border

    # ===== تنظیم عرض ستون‌ها =====
    column_widths = {
        'A': 6,  # ردیف
        'B': 18,  # شماره سفارش
        'C': 20,  # مشتری
        'D': 15,  # شماره موبایل
        'E': 35,  # آدرس
        'F': 14,  # تعداد محصولات
        'G': 18,  # مبلغ کل
        'H': 16,  # وضعیت سفارش
        'I': 16,  # وضعیت پرداخت
        'J': 18,  # تاریخ ثبت
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # ===== فریز کردن ردیف اول =====
    worksheet.freeze_panes = 'A2'

    return workbook